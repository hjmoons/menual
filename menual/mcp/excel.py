# -*- coding: utf-8 -*-
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


COLOR_BLUE      = "2E75B6"
COLOR_GRAY      = "A9A9A9"
COLOR_SEPARATOR = "D9D9D9"  # 원복 프로세스 구분선
COLOR_WHITE     = "FFFFFF"
COLOR_BLACK     = "000000"

# 6개 고정 step 정의
FIXED_STEPS = [
    {"no": 1, "label": "사전작업",                       "fixed": False},
    {"no": 2, "label": "작업개시선언",                   "fixed": True},   # 내용/시간 고정
    {"no": 3, "label": "본작업",                         "fixed": False},
    {"no": 4, "label": "서비스 오픈 선언 및 작업 종료",  "fixed": True},   # 내용/시간 고정
    # ── 원복 프로세스 구분선 (2행) ──
    {"no": 5, "label": "원복작업",                       "fixed": False},
    {"no": 6, "label": "점검 및 보고",                   "fixed": False},
]

TOTAL_COLS = 7  # A~G


def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _title_font(size=10):
    return Font(name="맑은 고딕", bold=True, color=COLOR_WHITE, size=size)

def _header_font():
    return Font(name="맑은 고딕", bold=True, color=COLOR_WHITE, size=8)

def _data_font(bold=False):
    return Font(name="맑은 고딕", size=8, color=COLOR_BLACK, bold=bold)


def _set_cell(ws, row, col, value="", bg=COLOR_WHITE, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.border = _border()
    c.font = font or _data_font()
    c.alignment = align or _center()
    return c


def _merge_set(ws, r1, c1, r2, c2, value="", bg=COLOR_WHITE, font=None, align=None):
    """병합 후 스타일 적용 (border는 전체 셀에)"""
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    _set_cell(ws, r1, c1, value, bg=bg, font=font, align=align or _center())
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _border()


def _make_timetable_sheet(ws, steps_data: list):
    """
    steps_data: 6개 항목 리스트 (FIXED_STEPS 순서 대응)
      각 항목: {
        "start_time": "09:00",
        "duration": "30분",        # fixed step이면 무시하고 "-" 사용
        "tasks": [                 # fixed step이면 무시
          {"content": "...", "reporter": "...", "worker": "...", "scenario": "..."}
        ]
      }
    항목이 부족하면 빈 값으로 채움
    """
    ws.title = "작업계획서"

    # 열 너비
    col_widths = {"A": 8, "B": 12, "C": 10, "D": 40, "E": 12, "F": 12, "G": 45}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── 제목 row 1 ──
    _merge_set(ws, 1, 1, 1, TOTAL_COLS, "작업계획서",
               bg=COLOR_BLUE, font=_title_font(10))
    ws.row_dimensions[1].height = 22

    # ── 헤더 row 2~3 ──
    # row2: Step | 시간계획(B~C 병합) | 작업 및 점검내용 | 보고자 | 작업자 | 수행시나리오
    _merge_set(ws, 2, 2, 2, 3, "시간계획", bg=COLOR_GRAY, font=_header_font())
    for col, val in [(1, "Step"), (4, "작업 및 점검내용"), (5, "보고자"), (6, "작업자"), (7, "수행시나리오")]:
        _set_cell(ws, 2, col, val, bg=COLOR_GRAY, font=_header_font())

    # row3: 소헤더 (시작시간 / 작업시간)
    for col, val in [(1, ""), (2, "시작시간"), (3, "작업시간"), (4, ""), (5, ""), (6, ""), (7, "")]:
        _set_cell(ws, 3, col, val, bg=COLOR_GRAY, font=_header_font())

    # row2~3 병합: Step, 작업및점검내용, 보고자, 작업자, 수행시나리오
    for col in (1, 4, 5, 6, 7):
        _merge_set(ws, 2, col, 3, col, ws.cell(2, col).value,
                   bg=COLOR_GRAY, font=_header_font())

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    # ── 데이터 행 ──
    cur_row = 4

    for idx, step_def in enumerate(FIXED_STEPS):
        # 구분선: step 4 다음 (idx==3 이후, idx==4 전)
        if idx == 4:
            ws.row_dimensions[cur_row].height = 18
            _merge_set(ws, cur_row, 1, cur_row, TOTAL_COLS,
                       "원복 프로세스 진행",
                       bg=COLOR_SEPARATOR,
                       font=_data_font(bold=True),
                       align=_center())
            cur_row += 1

        # 이 step의 입력 데이터
        d = steps_data[idx] if idx < len(steps_data) else {}
        start_time = d.get("start_time", "")
        duration   = "-" if step_def["fixed"] else d.get("duration", "")
        tasks      = [] if step_def["fixed"] else d.get("tasks", [])

        # 전체 행 수 = 1(label행) + len(tasks)
        span = 1 + len(tasks)

        # 행 초기화
        for i in range(span):
            r = cur_row + i
            ws.row_dimensions[r].height = 25
            for col in range(1, TOTAL_COLS + 1):
                _set_cell(ws, r, col)

        # step no / 시작시간 / 작업시간 → 세로 병합
        _merge_set(ws, cur_row, 1, cur_row + span - 1, 1,
                   step_def["no"], align=_center())
        _merge_set(ws, cur_row, 2, cur_row + span - 1, 2,
                   start_time, align=_center())
        _merge_set(ws, cur_row, 3, cur_row + span - 1, 3,
                   duration, align=_center())

        # label 행
        _set_cell(ws, cur_row, 4, step_def["label"],
                  font=_data_font(bold=True), align=_left())

        # task 행들
        for t_idx, task in enumerate(tasks):
            r = cur_row + 1 + t_idx
            _set_cell(ws, r, 4, task.get("content", ""),   align=_left())
            _set_cell(ws, r, 5, task.get("reporter", ""),  align=_center())
            _set_cell(ws, r, 6, task.get("worker", ""),    align=_center())
            _set_cell(ws, r, 7, task.get("scenario", ""),  align=_left())

        cur_row += span


# ── 사전작업 / 본작업 / 원복작업 공통 시트 ───────────────────────────────────

TASK_COLS = [
    ("A", 8,  "작업연번"),
    ("B", 20, "작업내용"),
    ("C", 14, "대상"),
    ("D", 16, "대상장비"),
    ("E", 30, "세부 작업내용"),
    ("F", 16, "작업자/보고자"),
    ("G", 10, "완료여부"),
    ("H", 35, "command"),
    ("I", 20, "비고"),
]


def _make_sub_sheet(ws, title: str, groups: list):
    """
    groups: [
        {
            "content": "작업내용",
            "target": "대상",
            "device": "대상장비",
            "details": [
                {"detail": "세부 작업내용", "owner": "작업자/보고자", "done": "", "command": "명령어", "note": "비고"}
            ]
        }, ...
    ]
    """
    ws.title = title

    for col_letter, width, _ in TASK_COLS:
        ws.column_dimensions[col_letter].width = width

    # 제목
    ws.merge_cells(f"A1:{TASK_COLS[-1][0]}1")
    _set_cell(ws, 1, 1, title, bg=COLOR_BLUE, font=_title_font(10))
    ws.row_dimensions[1].height = 22

    # 헤더
    for col_idx, (_, _, label) in enumerate(TASK_COLS, start=1):
        _set_cell(ws, 2, col_idx, label, bg=COLOR_GRAY, font=_header_font())
    ws.row_dimensions[2].height = 18

    ws.freeze_panes = "A3"

    if not groups:
        for i in range(20):
            r = 3 + i
            ws.row_dimensions[r].height = 25
            for col_idx in range(1, len(TASK_COLS) + 1):
                _set_cell(ws, r, col_idx, align=_center() if col_idx in (1, 6, 7) else _left())
            ws.cell(r, 1).value = i + 1
        return

    cur_row = 3
    for i, group in enumerate(groups):
        details = group.get("details", [{}])
        span = len(details)

        for d_idx in range(span):
            r = cur_row + d_idx
            ws.row_dimensions[r].height = 25
            for col_idx in range(1, len(TASK_COLS) + 1):
                _set_cell(ws, r, col_idx, align=_center() if col_idx in (1, 6, 7) else _left())

        # 연번 / 작업내용 / 대상 / 대상장비 → 세로 병합
        _merge_set(ws, cur_row, 1, cur_row + span - 1, 1, i + 1, align=_center())
        _merge_set(ws, cur_row, 2, cur_row + span - 1, 2, group.get("content", ""), align=_left())
        _merge_set(ws, cur_row, 3, cur_row + span - 1, 3, group.get("target", ""), align=_center())
        _merge_set(ws, cur_row, 4, cur_row + span - 1, 4, group.get("device", ""), align=_center())

        for d_idx, detail in enumerate(details):
            r = cur_row + d_idx
            _set_cell(ws, r, 5, detail.get("detail", ""),   align=_left())
            _set_cell(ws, r, 6, detail.get("owner", ""),    align=_center())
            _set_cell(ws, r, 7, detail.get("done", ""),     align=_center())
            _set_cell(ws, r, 8, detail.get("command", ""),  align=_left())
            _set_cell(ws, r, 9, detail.get("note", ""),     align=_left())

        cur_row += span


# ── Workbook ──────────────────────────────────────────────────────────────────

def _make_workbook(task_name: str, steps_data: list,
                   pre_rows=None, main_rows=None, rollback_rows=None) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("작업계획서")
    _make_timetable_sheet(ws1, steps_data)

    ws2 = wb.create_sheet("사전작업")
    _make_sub_sheet(ws2, "사전작업", pre_rows or [])

    ws3 = wb.create_sheet("본작업")
    _make_sub_sheet(ws3, "본작업", main_rows or [])

    ws4 = wb.create_sheet("원복작업")
    _make_sub_sheet(ws4, "원복작업", rollback_rows or [])

    return wb


def register(mcp):

    @mcp.tool()
    def create_work_plan_excel(
        task_name: str,
        steps: str = "[]",
        pre_tasks: str = "[]",
        main_tasks: str = "[]",
        rollback_tasks: str = "[]",
        save_dir: str = ""
    ) -> str:
        """엑셀 작업계획서 생성 (작업계획서/사전작업/본작업/원복작업 4개 시트)

        Args:
            task_name: 작업 이름
            steps: 작업계획서 시트용. 6개 step 데이터 JSON 배열 (순서: 사전작업/작업개시선언/본작업/서비스오픈선언/원복작업/점검및보고)
                각 항목: {"start_time": "09:00", "duration": "30분", "tasks": [{"content": "...", "reporter": "...", "worker": "...", "scenario": "..."}]}
                작업개시선언(2번), 서비스오픈선언(4번)은 start_time만 사용하고 duration/tasks는 무시됨
                scenario 필드는 명령어가 아닌 보고 메시지나 진행 상황을 자연어로 기술할 것.
                예) "작업 시작을 팀장님께 메신저로 보고", "@팀장 nginx 재시작 완료, 서비스 정상 확인"
            pre_tasks: 사전작업 시트 데이터 JSON 배열
                각 항목: {"content": "작업내용", "target": "대상", "device": "대상장비",
                          "details": [{"detail": "세부작업내용", "owner": "작업자/보고자", "done": "", "command": "명령어", "note": "비고"}]}
                작업내용/대상/대상장비는 details 수만큼 셀이 세로 병합됨
            main_tasks: 본작업 시트 데이터 JSON 배열 (pre_tasks와 동일 구조)
            rollback_tasks: 원복작업 시트 데이터 JSON 배열 (pre_tasks와 동일 구조)
            save_dir: 저장 폴더 경로 (비우면 바탕화면)
        """
        try:
            steps_data    = json.loads(steps)
            pre_rows      = json.loads(pre_tasks)
            main_rows     = json.loads(main_tasks)
            rollback_rows = json.loads(rollback_tasks)
        except json.JSONDecodeError as e:
            return f"JSON 파싱 오류: {e}"

        out_dir = Path(save_dir) if save_dir else Path.home() / "Desktop"
        out_dir.mkdir(parents=True, exist_ok=True)

        date_str  = datetime.now().strftime("%Y%m%d")
        safe_name = task_name.replace(" ", "_").replace("/", "-")
        out_path  = out_dir / f"작업계획서_{safe_name}_{date_str}.xlsx"

        _make_workbook(task_name, steps_data, pre_rows, main_rows, rollback_rows).save(out_path)
        return f"저장 완료: {out_path}"
